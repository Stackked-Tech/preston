#!/usr/bin/env python3
"""
Phorest CSV → NetSuite Payroll XLSX Transformer

Fetches transaction data from the Phorest API and generates a NetSuite-ready
payroll import spreadsheet for William Henry Salon.

Usage:
    # Fetch from Phorest API and generate NetSuite xlsx:
    python3 phorest_to_netsuite.py fetch --start 2026-02-02 --end 2026-02-14 --pay-date 2026-02-19

    # Process a local CSV file:
    python3 phorest_to_netsuite.py local <input_csv> [--output <output_xlsx>]

Environment Variables (in .env):
    PHOREST_API_URL, PHOREST_BUSINESS_ID, PHOREST_BRANCH_ID,
    PHOREST_USERNAME, PHOREST_PASSWORD
"""

import csv
import sys
import os
import json
import time
import tempfile
from datetime import datetime, date, timedelta
from collections import defaultdict
from urllib.request import Request, urlopen
from urllib.error import HTTPError, URLError
from base64 import b64encode

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip3 install openpyxl")
    sys.exit(1)


# ═══════════════════════════════════════════════════════════════════════════════
# ENV / CREDENTIALS
# ═══════════════════════════════════════════════════════════════════════════════

def load_env():
    """Load .env file from project root."""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    # Look for .env in parent directories
    for parent in [script_dir, os.path.dirname(script_dir)]:
        env_path = os.path.join(parent, '.env')
        if os.path.exists(env_path):
            with open(env_path, 'r') as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith('#') and '=' in line:
                        key, val = line.split('=', 1)
                        os.environ.setdefault(key.strip(), val.strip())
            return
    print("WARNING: No .env file found")


def get_phorest_config():
    """Get Phorest API configuration from environment."""
    required = ['PHOREST_API_URL', 'PHOREST_BUSINESS_ID', 'PHOREST_BRANCH_ID',
                'PHOREST_USERNAME', 'PHOREST_PASSWORD']
    config = {}
    missing = []
    for key in required:
        val = os.environ.get(key)
        if not val:
            missing.append(key)
        config[key] = val

    if missing:
        print(f"ERROR: Missing environment variables: {', '.join(missing)}")
        print("Set them in .env or export them in your shell.")
        sys.exit(1)

    return config


# ═══════════════════════════════════════════════════════════════════════════════
# PHOREST API
# ═══════════════════════════════════════════════════════════════════════════════

def phorest_request(method, path, body=None):
    """Make an authenticated request to the Phorest API."""
    env = get_phorest_config()
    url = f"{env['PHOREST_API_URL']}{path}"
    credentials = b64encode(
        f"{env['PHOREST_USERNAME']}:{env['PHOREST_PASSWORD']}".encode()
    ).decode()

    headers = {
        'Authorization': f'Basic {credentials}',
        'Content-Type': 'application/json',
    }

    data = json.dumps(body).encode() if body else None
    req = Request(url, data=data, headers=headers, method=method)

    try:
        with urlopen(req) as resp:
            return json.loads(resp.read().decode())
    except HTTPError as e:
        error_body = e.read().decode()
        print(f"ERROR: Phorest API returned {e.code}")
        try:
            print(f"  {json.loads(error_body).get('detail', error_body)}")
        except json.JSONDecodeError:
            print(f"  {error_body[:200]}")
        sys.exit(1)


def fetch_csv_from_phorest(start_date, end_date):
    """Create a CSV export job, poll until done, download the CSV."""
    env = get_phorest_config()
    biz = env['PHOREST_BUSINESS_ID']
    branch = env['PHOREST_BRANCH_ID']
    base_path = f"/api/business/{biz}/branch/{branch}/csvexportjob"

    # 1. Create the export job
    print(f"Creating Phorest CSV export job...")
    print(f"  Date range: {start_date} to {end_date}")
    job = phorest_request('POST', base_path, {
        'startFilter': start_date.isoformat(),
        'finishFilter': end_date.isoformat(),
        'jobType': 'TRANSACTIONS_CSV',
    })

    job_id = job['jobId']
    print(f"  Job ID: {job_id}")
    print(f"  Status: {job['jobStatus']}")

    # 2. Poll until DONE
    poll_interval = 5  # seconds
    max_wait = 600     # 10 minutes
    elapsed = 0

    while elapsed < max_wait:
        time.sleep(poll_interval)
        elapsed += poll_interval

        status = phorest_request('GET', f"{base_path}/{job_id}")
        job_status = status.get('jobStatus', 'UNKNOWN')
        rows_done = status.get('succeededRows', '?')
        rows_total = status.get('totalRows', '?')

        print(f"  [{elapsed}s] Status: {job_status}  "
              f"Rows: {rows_done}/{rows_total}", end='\r')

        if job_status == 'DONE':
            print(f"  [{elapsed}s] Status: DONE  "
                  f"Rows: {rows_done}/{rows_total}       ")
            break
    else:
        print(f"\nERROR: Job timed out after {max_wait}s")
        sys.exit(1)

    # 3. Check for failure
    failure = status.get('failureReason')
    if failure:
        print(f"  Job failed: {failure}")
        sys.exit(1)

    download_url = status.get('tempCsvExternalUrl')
    if not download_url:
        print("  ERROR: No download URL in completed job")
        sys.exit(1)

    # 4. Download the CSV
    print(f"  Downloading CSV ({rows_total} rows)...")
    req = Request(download_url)
    with urlopen(req) as resp:
        csv_data = resp.read()

    # Save to a temp file
    script_dir = os.path.dirname(os.path.abspath(__file__))
    csv_path = os.path.join(
        script_dir,
        f"phorest_export_{start_date.isoformat()}_{end_date.isoformat()}.csv"
    )
    with open(csv_path, 'wb') as f:
        f.write(csv_data)

    print(f"  Saved to: {csv_path}")
    return csv_path


# ═══════════════════════════════════════════════════════════════════════════════
# DATE COMPUTATION
# ═══════════════════════════════════════════════════════════════════════════════

def compute_pay_period_config(period_start, period_end, pay_date):
    """Compute all derived date values from the three input dates."""

    # Week 1 ends on the first Saturday on or after the start date
    # (pay periods start on Monday, each "week" is Mon-Sat)
    d = period_start
    while d.weekday() != 5:  # 5 = Saturday
        d += timedelta(days=1)
    week1_end = d

    # Posting period = 1st of the month
    posting_period = date(period_start.year, period_start.month, 1)

    # Pay period label: "{M}/{label_start}-{end_day}/{Y}"
    # First half of month: label starts at 1, second half: 16
    if period_start.day <= 15:
        label_start = 1
    else:
        label_start = 16
    pay_period_label = (f"{period_start.month}/{label_start}-"
                        f"{period_end.day}/{period_end.year}")

    return {
        "period_start": period_start,
        "period_end": period_end,
        "week1_end": week1_end,
        "pay_date": pay_date,
        "posting_period": posting_period,
        "pay_period_label": pay_period_label,
        "subsidiary_id": 5,
        "account": 111,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# NEW GUEST QUALIFYING CLIENT SOURCES
# ═══════════════════════════════════════════════════════════════════════════════

NEW_GUEST_SOURCES = {
    "Call In ---New Guest",
    "Online Booking ---New Guest",
    "Walk In --New Guest",
}


# ═══════════════════════════════════════════════════════════════════════════════
# STAFF CONFIGURATION
# Key = Phorest name (first last as it appears in the CSV)
#
# Required fields: target_first, target_last, internal_id, station_lease,
#                  financial_services, phorest_fee, refreshment
# Optional fields: associate_pay (manual entry for L column),
#                  supervisor (Phorest name of supervisor for associate fee)
# ═══════════════════════════════════════════════════════════════════════════════

STAFF_CONFIG = {
    "Danielle Baker": {
        "target_first": "Danielle", "target_last": "Seeger Baker",
        "internal_id": 1736, "station_lease": -390, "financial_services": -100,
        "phorest_fee": -10, "refreshment": -10,
    },
    "Gabby Brewer": {
        "target_first": "Gabby", "target_last": "Brewer",
        "internal_id": 3072, "station_lease": 0, "financial_services": 0,
        "phorest_fee": 0, "refreshment": 0,
    },
    "Emma Davis": {
        "target_first": "Emma", "target_last": "Baldwin-Davis",
        "internal_id": 1738, "station_lease": -336, "financial_services": -100,
        "phorest_fee": -10, "refreshment": -10,
    },
    "Addison Brown": {
        "target_first": "Addison", "target_last": "Brown",
        "internal_id": 3071, "station_lease": 0, "financial_services": -100,
        "phorest_fee": -10, "refreshment": -10,
        "associate_pay": None,  # Manual entry — set per pay period
        "supervisor": "Seth King",
    },
    "Grace Deason": {
        "target_first": "Grace", "target_last": "Deason",
        "internal_id": 1743, "station_lease": -336, "financial_services": -100,
        "phorest_fee": -10, "refreshment": -10,
    },
    "Molly Diaz": {
        "target_first": "Molly", "target_last": "Diaz",
        "internal_id": 1758, "station_lease": -336, "financial_services": -100,
        "phorest_fee": -10, "refreshment": -10,
    },
    "Ashleigh Dotson": {
        "target_first": "Ashleigh", "target_last": "Dotson",
        "internal_id": 1726, "station_lease": -390, "financial_services": -100,
        "phorest_fee": -10, "refreshment": -10,
    },
    "Kristen Forehand": {
        "target_first": "Kristen", "target_last": "Forehand",
        "internal_id": 1752, "station_lease": -238, "financial_services": -100,
        "phorest_fee": -10, "refreshment": -10,
    },
    "Aubrey Hawkins": {
        "target_first": "Aubrey", "target_last": "Hawkins",
        "internal_id": 2667, "station_lease": -25, "financial_services": -50,
        "phorest_fee": -5, "refreshment": -5,
    },
    "Jess Herzog": {
        "target_first": "Jess", "target_last": "Herzog",
        "internal_id": 2296, "station_lease": -336, "financial_services": -100,
        "phorest_fee": -10, "refreshment": -10,
    },
    "Kaylie Houghtaling": {
        "target_first": "Kaylie", "target_last": "Houghtaling",
        "internal_id": 1751, "station_lease": -336, "financial_services": -100,
        "phorest_fee": -10, "refreshment": -10,
    },
    "Seth King": {
        "target_first": "Seth ", "target_last": "King",
        "internal_id": 1771, "station_lease": -336, "financial_services": -100,
        "phorest_fee": -10, "refreshment": -10,
    },
    "Grace Lesser": {
        "target_first": "Grace", "target_last": "Lesser",
        "internal_id": 3074, "station_lease": 0, "financial_services": 0,
        "phorest_fee": 0, "refreshment": 0,
        "associate_pay": None,
        "supervisor": None,
    },
    "Cassi Mcclure": {
        "target_first": "Cassi", "target_last": "McClure",
        "internal_id": 1731, "station_lease": -390, "financial_services": -100,
        "phorest_fee": -10, "refreshment": -10,
    },
    "Brooke Parker": {
        "target_first": "Brooke", "target_last": "Parker",
        "internal_id": 2671, "station_lease": -320, "financial_services": -100,
        "phorest_fee": -10, "refreshment": -10,
    },
    "Keleigh Ratliff": {
        "target_first": "Keleigh", "target_last": "Ratliff",
        "internal_id": 1635, "station_lease": 0, "financial_services": 0,
        "phorest_fee": 0, "refreshment": 0,
    },
    "Torey Rome": {
        "target_first": "Torey", "target_last": "Rome",
        "internal_id": 1775, "station_lease": -390, "financial_services": -100,
        "phorest_fee": -10, "refreshment": -10,
    },
    "Maddie Schultz": {
        "target_first": "Maddie", "target_last": "Schultz",
        "internal_id": 3073, "station_lease": 0, "financial_services": 0,
        "phorest_fee": 0, "refreshment": 0,
        "associate_pay": None,
        "supervisor": None,
    },
    "Sierra Sharpe": {
        "target_first": "Sierra", "target_last": "Sharpe",
        "internal_id": 3077, "station_lease": 0, "financial_services": -50,
        "phorest_fee": -5, "refreshment": -5,
        "associate_pay": None,
        "supervisor": "Kristen Forehand",
    },
    "Dana Siepert": {
        "target_first": "Dana", "target_last": "Siepert",
        "internal_id": 2295, "station_lease": -336, "financial_services": -100,
        "phorest_fee": -10, "refreshment": -10,
    },
    "Lauren Simonds": {
        "target_first": "Lauren", "target_last": "Simonds",
        "internal_id": 1754, "station_lease": -336, "financial_services": -100,
        "phorest_fee": -10, "refreshment": -10,
    },
    "Aubrey White": {
        "target_first": "Aubrey", "target_last": "White",
        "internal_id": 2670, "station_lease": -320, "financial_services": -100,
        "phorest_fee": -10, "refreshment": -10,
    },
    "Olivia Wilson": {
        "target_first": "Olivia", "target_last": "Cornette Wilson",
        "internal_id": 1765, "station_lease": -390, "financial_services": -100,
        "phorest_fee": -10, "refreshment": -10,
    },
}

# Employee Purchase client name overrides (client name → Phorest staff name)
# Used when the client name on an ES transaction doesn't directly match a Phorest name
EMPLOYEE_PURCHASE_NAME_MAP = {
    "Olivia Cornette": "Olivia Wilson",
}

# Display order in the output spreadsheet
STAFF_ORDER = [
    "Danielle Baker", "Gabby Brewer", "Emma Davis", "Addison Brown",
    "Grace Deason", "Molly Diaz", "Ashleigh Dotson", "Kristen Forehand",
    "Aubrey Hawkins", "Jess Herzog", "Kaylie Houghtaling", "Seth King",
    "Grace Lesser", "Cassi Mcclure", "Brooke Parker", "Keleigh Ratliff",
    "Torey Rome", "Maddie Schultz", "Sierra Sharpe", "Dana Siepert",
    "Lauren Simonds", "Aubrey White", "Olivia Wilson",
]


# ═══════════════════════════════════════════════════════════════════════════════
# PROCESSING LOGIC
# ═══════════════════════════════════════════════════════════════════════════════

def parse_date(date_str):
    """Parse various date formats from the CSV."""
    if not date_str or date_str.strip() == '':
        return None
    date_str = date_str.strip()
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%m/%d/%Y', '%m/%d/%y'):
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    return None


def process_csv(csv_path, config):
    """Process the Phorest CSV and return computed values per staff member."""

    results = {name: {
        'product_wk1': 0.0,
        'product_wk2': 0.0,
        'service_total': 0.0,
        'tips': 0.0,
        'new_guests': 0.0,
        'employee_purchases': 0.0,
    } for name in STAFF_CONFIG}

    # Deduplication tracking
    seen_tips = set()   # (staff_name, transaction_id)
    seen_es = set()     # (client_name, transaction_id)

    # Employee purchase amounts by client name (resolve to staff after)
    emp_purchases_by_client = defaultdict(float)

    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            staff_first = row.get('staff_first_name', '').strip()
            staff_last = row.get('staff_last_name', '').strip()
            staff_name = f"{staff_first} {staff_last}"
            item_type = row.get('item_type', '').strip()
            transaction_id = row.get('transaction_id', '').strip()
            purchased_date = parse_date(row.get('purchased_date', ''))

            is_known_staff = staff_name in results

            # ── Product Sales (wk1/wk2) ──
            if item_type == 'PRODUCT' and is_known_staff and purchased_date:
                total = float(row.get('total_amount', 0) or 0)
                if purchased_date <= config['week1_end']:
                    results[staff_name]['product_wk1'] += total
                else:
                    results[staff_name]['product_wk2'] += total

            # ── Service Totals (Col K) ──
            if item_type == 'SERVICE' and is_known_staff:
                total = float(row.get('total_amount', 0) or 0)
                results[staff_name]['service_total'] += total

            # ── Tips (only GC transactions, deduplicate by transaction_id) ──
            # Tips from Gift Card payments need separate payout (CC tips go
            # through CC processor). Additional cash tips may need manual entry.
            if is_known_staff and transaction_id:
                tip_key = (staff_name, transaction_id)
                if tip_key not in seen_tips:
                    seen_tips.add(tip_key)
                    tips = float(row.get('phorest_tips', 0) or 0)
                    if tips > 0:
                        codes = row.get('payment_type_codes', '').strip()
                        code_list = [c.strip() for c in codes.split(';')]
                        if 'GC' in code_list:
                            results[staff_name]['tips'] += tips

            # ── New Guests ──
            if item_type == 'SERVICE' and is_known_staff:
                client_source = row.get('client_source', '').strip()
                client_first_visit = parse_date(row.get('client_first_visit', ''))
                if (client_source in NEW_GUEST_SOURCES and
                        client_first_visit and
                        config['period_start'] <= client_first_visit <= config['period_end']):
                    unit_price = float(row.get('unit_price', 0) or 0)
                    results[staff_name]['new_guests'] += unit_price

            # ── Employee Purchases ──
            payment_names = row.get('payment_type_names', '').strip()
            if 'Employee Sale' in payment_names and transaction_id:
                client_first = row.get('client_first_name', '').strip()
                client_last = row.get('client_last_name', '').strip()
                client_name = f"{client_first} {client_last}"
                es_key = (client_name, transaction_id)
                if es_key not in seen_es:
                    seen_es.add(es_key)
                    codes = row.get('payment_type_codes', '').split(';')
                    amounts = row.get('payment_type_amounts', '').split(';')
                    for code, amount in zip(codes, amounts):
                        if code.strip() == 'ES':
                            emp_purchases_by_client[client_name] += float(amount.strip() or 0)

    # ── Resolve employee purchase client names to staff names ──
    for client_name, amount in emp_purchases_by_client.items():
        # 1) Check explicit override map
        if client_name in EMPLOYEE_PURCHASE_NAME_MAP:
            staff = EMPLOYEE_PURCHASE_NAME_MAP[client_name]
            results[staff]['employee_purchases'] += amount
            continue

        # 2) Direct match to Phorest name
        if client_name in results:
            results[client_name]['employee_purchases'] += amount
            continue

        # 3) First-name match + client last name in target_last
        matched = False
        client_parts = client_name.split()
        if client_parts:
            client_first_lower = client_parts[0].lower()
            client_last_part = ' '.join(client_parts[1:]).lower() if len(client_parts) > 1 else ''
            for phorest_name, cfg in STAFF_CONFIG.items():
                phorest_first = phorest_name.split()[0].lower()
                if (phorest_first == client_first_lower and
                        client_last_part and
                        client_last_part in cfg['target_last'].lower()):
                    results[phorest_name]['employee_purchases'] += amount
                    matched = True
                    break

        if not matched:
            print(f"  WARNING: Could not match employee purchase client "
                  f"'{client_name}' to any staff member (amount: ${amount:.2f})")

    return results


def generate_xlsx(results, config, output_path):
    """Generate the NetSuite payroll import XLSX."""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # ── Header Row 1 ──
    headers_r1 = [
        '', '', '', '', 'Product Sales', '', 'Product Sales', 'Booth Rent',
        '(GL 7140)', '(GL 7150)', '(GL 7170)', '(GL 7180)', 'Total',
        '(GL 4020)', '(GL 4030)', '(GL 4045)', 'Credit Card Amount',
        '(GL 4040)', 'New Guests', '(GL 4060)', '(GL 4070)', '(GL 4080)',
        '(GL 4090)', '(GL 4120)', 'Misc. Fees', 'Total', 'Account',
        'Posting Period', 'Reference #', 'Due Date',
        'Approval Status\n(Approved/\nPending)', '',
    ]

    # ── Header Row 2 ──
    headers_r2 = [
        '', '', '', '', '(wk 1)', 'Booth Rent', '(wk 2)', 'Rebate (wk 2)',
        'Booth Rent', 'Tips', 'Contractor', 'Associate', 'Earned',
        'Station', 'Financial', 'Color', '', 'Credit Card', '', 'Finders',
        'Employee', 'Phorest', 'Refreshment', 'Associate', '', 'check',
        '', '', '', '', '', '',
    ]

    # ── Header Row 3 ──
    headers_r3 = [
        'Subsidiary ID', 'Internal ID', 'First Names', 'Last Name', '',
        'Rebate (wk 1)', '', '', 'Rebate Total', '', 'Service', 'Pay',
        '', 'Lease', 'Services', 'Charges', '', 'Charges 3%', '', 'Fee 20%',
        'Purchases', '', '', 'Fee', '', '', '', '', '', '', '', 'Pay Period',
    ]

    for col_idx, val in enumerate(headers_r1, 1):
        ws.cell(row=1, column=col_idx, value=val if val else None)
    for col_idx, val in enumerate(headers_r2, 1):
        ws.cell(row=2, column=col_idx, value=val if val else None)
    for col_idx, val in enumerate(headers_r3, 1):
        ws.cell(row=3, column=col_idx, value=val if val else None)

    # ── Data Rows ──
    for i, phorest_name in enumerate(STAFF_ORDER):
        row = 4 + i
        cfg = STAFF_CONFIG[phorest_name]
        data = results.get(phorest_name, {})

        # A: Subsidiary ID
        ws.cell(row=row, column=1, value=config['subsidiary_id'])

        # B: Internal ID
        ws.cell(row=row, column=2, value=cfg['internal_id'])

        # C: First Name
        ws.cell(row=row, column=3, value=cfg['target_first'])

        # D: Last Name
        ws.cell(row=row, column=4, value=cfg['target_last'])

        # E: Product Sales (wk 1)
        prod_wk1 = round(data.get('product_wk1', 0), 2)
        ws.cell(row=row, column=5, value=prod_wk1 if prod_wk1 else None)

        # F: Booth Rent Rebate (wk 1) — FORMULA
        ws.cell(row=row, column=6,
                value=f'=IF(E{row}>250,0.2,IF(E{row}>149,0.15,IF(E{row}>49,0.1,0)))*E{row}')

        # G: Product Sales (wk 2)
        prod_wk2 = round(data.get('product_wk2', 0), 2)
        ws.cell(row=row, column=7, value=prod_wk2 if prod_wk2 else None)

        # H: Booth Rent Rebate (wk 2) — FORMULA
        ws.cell(row=row, column=8,
                value=f'=IF(G{row}>250,0.2,IF(G{row}>149,0.15,IF(G{row}>49,0.1,0)))*G{row}')

        # I: Booth Rent Rebate Total — FORMULA
        ws.cell(row=row, column=9, value=f'=F{row}+H{row}')

        # J: Tips
        tips = round(data.get('tips', 0), 2)
        ws.cell(row=row, column=10, value=tips if tips else None)

        # K: Contractor Service (all staff service totals)
        svc = round(data.get('service_total', 0), 2)
        ws.cell(row=row, column=11, value=svc if svc else None)

        # L: Associate Pay (manual entry)
        assoc_pay = cfg.get('associate_pay')
        ws.cell(row=row, column=12, value=assoc_pay)

        # M: Total Earned — FORMULA
        ws.cell(row=row, column=13, value=f'=SUM(I{row}+J{row}+K{row}+L{row})')

        # N: Station Lease
        ws.cell(row=row, column=14, value=cfg['station_lease'])

        # O: Financial Services
        ws.cell(row=row, column=15, value=cfg['financial_services'])

        # P: Color Charges (blank — manual entry)
        ws.cell(row=row, column=16, value=None)

        # Q: Credit Card Amount (blank — manual entry)
        ws.cell(row=row, column=17, value=None)

        # R: Credit Card Charges 3% — FORMULA
        ws.cell(row=row, column=18, value=f'=0.03*(-Q{row})')

        # S: New Guests
        new_guests = round(data.get('new_guests', 0), 2)
        ws.cell(row=row, column=19, value=new_guests if new_guests else None)

        # T: Finders Fee 20% — FORMULA
        ws.cell(row=row, column=20, value=f'=0.2*(-S{row})')

        # U: Employee Purchases
        emp_purch = round(data.get('employee_purchases', 0), 2)
        ws.cell(row=row, column=21, value=-emp_purch if emp_purch else None)

        # V: Phorest
        ws.cell(row=row, column=22, value=cfg['phorest_fee'])

        # W: Refreshment
        ws.cell(row=row, column=23, value=cfg['refreshment'])

        # X: Associate Fee (computed from associate_pay if supervisor mapping exists)
        # This is populated on the SUPERVISOR's row, not the associate's row
        ws.cell(row=row, column=24, value=None)

        # Y: Misc Fees (blank)
        ws.cell(row=row, column=25, value=None)

        # Z: Total Check — FORMULA
        ws.cell(row=row, column=26,
                value=f'=M{row}+(N{row}+O{row}+P{row}+R{row}+T{row}+U{row}+V{row}+W{row}+X{row}+Y{row})')

        # AA: Account
        ws.cell(row=row, column=27, value=config['account'])

        # AB: Posting Period
        ws.cell(row=row, column=28, value=config['posting_period'])

        # AC: Reference #
        pd = config['pay_date']
        ws.cell(row=row, column=29, value=f"ACH {pd.month}.{pd.day}.{pd.year}")

        # AD: Due Date
        ws.cell(row=row, column=30, value=config['pay_date'])

        # AE: Approval Status
        ws.cell(row=row, column=31, value="Approved")

        # AF: Pay Period
        ws.cell(row=row, column=32, value=config['pay_period_label'])

    # ── Associate Fee (Col X) — populated on supervisor rows ──
    for phorest_name, cfg in STAFF_CONFIG.items():
        supervisor = cfg.get('supervisor')
        assoc_pay = cfg.get('associate_pay')
        if supervisor and assoc_pay and supervisor in STAFF_CONFIG:
            if supervisor in STAFF_ORDER:
                sup_row = 4 + STAFF_ORDER.index(supervisor)
                existing = ws.cell(row=sup_row, column=24).value
                new_val = (existing or 0) + (-assoc_pay)
                ws.cell(row=sup_row, column=24, value=new_val)

    # ── Footer Rows ──
    last_data_row = 3 + len(STAFF_ORDER)
    footer_r1 = last_data_row + 1  # TOTAL PAYROLL
    footer_r2 = last_data_row + 2  # TOTAL EMP. W/DRAWL
    footer_r3 = last_data_row + 3  # TOTAL ACH

    ws.cell(row=footer_r1, column=3, value="WHS 02")
    ws.cell(row=footer_r1, column=25, value="TOTAL PAYROLL:")
    ws.cell(row=footer_r1, column=26, value=f'=SUM(Z4:Z{last_data_row})')

    ws.cell(row=footer_r2, column=3, value="Pay Period:")
    ws.cell(row=footer_r2, column=4, value=config['pay_period_label'])
    ws.cell(row=footer_r2, column=25, value="TOTAL EMP. W/DRAWL:")

    ws.cell(row=footer_r3, column=3, value="Pay Date:")
    ws.cell(row=footer_r3, column=4, value=config['pay_date'])
    ws.cell(row=footer_r3, column=25, value="TOTAL ACH:")
    ws.cell(row=footer_r3, column=26, value=f'=Z{footer_r1}+Z{footer_r2}')

    wb.save(output_path)
    return output_path


def print_summary(results):
    """Print a summary comparison of computed values."""
    print("\n" + "=" * 80)
    print("COMPUTED VALUES SUMMARY")
    print("=" * 80)
    print(f"{'Staff':<25s} {'Prod W1':>8s} {'Prod W2':>8s} {'Service':>9s} "
          f"{'Tips':>7s} {'NewGuest':>9s} {'EmpPurch':>9s}")
    print("-" * 80)
    for name in STAFF_ORDER:
        d = results[name]
        pw1 = f"{d['product_wk1']:.2f}" if d['product_wk1'] else ""
        pw2 = f"{d['product_wk2']:.2f}" if d['product_wk2'] else ""
        svc = f"{d['service_total']:.2f}" if d['service_total'] else ""
        tips = f"{d['tips']:.2f}" if d['tips'] else ""
        ng = f"{d['new_guests']:.2f}" if d['new_guests'] else ""
        ep = f"{-d['employee_purchases']:.2f}" if d['employee_purchases'] else ""
        print(f"{name:<25s} {pw1:>8s} {pw2:>8s} {svc:>9s} {tips:>7s} {ng:>9s} {ep:>9s}")
    print("=" * 80)


def print_usage():
    print("""
Phorest → NetSuite Payroll Transformer

Usage:
  Fetch from Phorest API:
    python3 phorest_to_netsuite.py fetch --start YYYY-MM-DD --end YYYY-MM-DD --pay-date YYYY-MM-DD

  Process local CSV:
    python3 phorest_to_netsuite.py local <input_csv> --start YYYY-MM-DD --end YYYY-MM-DD --pay-date YYYY-MM-DD [--output <output.xlsx>]

Required flags:
  --start      First business day of the pay period (e.g., 2026-02-02)
  --end        Last business day of the pay period (e.g., 2026-02-14)
  --pay-date   Payment date / due date (e.g., 2026-02-19)

Optional:
  --output     Output xlsx file path (auto-generated if not specified)
""")


def parse_args():
    """Parse command-line arguments."""
    args = sys.argv[1:]
    if not args or args[0] in ('-h', '--help'):
        print_usage()
        sys.exit(0)

    mode = args[0]
    if mode not in ('fetch', 'local'):
        print_usage()
        sys.exit(1)

    def get_flag(flag):
        if flag in args:
            idx = args.index(flag)
            if idx + 1 < len(args):
                return args[idx + 1]
        return None

    start = get_flag('--start')
    end = get_flag('--end')
    pay_date = get_flag('--pay-date')
    output = get_flag('--output')

    if not start or not end or not pay_date:
        print("ERROR: --start, --end, and --pay-date are all required.")
        print_usage()
        sys.exit(1)

    try:
        start_date = date.fromisoformat(start)
        end_date = date.fromisoformat(end)
        pay_date_parsed = date.fromisoformat(pay_date)
    except ValueError as e:
        print(f"ERROR: Invalid date format: {e}")
        print("Use YYYY-MM-DD format.")
        sys.exit(1)

    csv_path = None
    if mode == 'local':
        # Find the CSV path (first arg that's not a flag)
        for i, a in enumerate(args[1:], 1):
            if not a.startswith('--') and (i == 1 or not args[i - 1].startswith('--')):
                csv_path = a
                break
        if not csv_path:
            print("ERROR: local mode requires a CSV file path.")
            print_usage()
            sys.exit(1)

    return {
        'mode': mode,
        'start_date': start_date,
        'end_date': end_date,
        'pay_date': pay_date_parsed,
        'csv_path': csv_path,
        'output': output,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    load_env()
    parsed = parse_args()

    # Compute pay period config from dates
    config = compute_pay_period_config(
        parsed['start_date'], parsed['end_date'], parsed['pay_date']
    )

    print(f"Pay Period: {config['pay_period_label']}")
    print(f"  Period:  {config['period_start']} to {config['period_end']}")
    print(f"  Week 1:  {config['period_start']} to {config['week1_end']}")
    print(f"  Week 2:  {config['week1_end'] + timedelta(days=1)} to {config['period_end']}")
    print(f"  Pay Date: {config['pay_date']}")
    print()

    # Get the CSV
    if parsed['mode'] == 'fetch':
        csv_path = fetch_csv_from_phorest(parsed['start_date'], parsed['end_date'])
    else:
        csv_path = parsed['csv_path']
        if not os.path.exists(csv_path):
            print(f"ERROR: Input file not found: {csv_path}")
            sys.exit(1)

    # Process
    print(f"\nProcessing: {csv_path}")
    results = process_csv(csv_path, config)
    print_summary(results)

    # Generate output
    output_path = parsed['output']
    if not output_path:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        label = config['pay_period_label'].replace('/', '-')
        output_path = os.path.join(script_dir, f"NetSuite_Payroll_{label}.xlsx")

    output = generate_xlsx(results, config, output_path)
    print(f"\nOutput saved to: {output}")
    print("\nMANUAL ENTRY REQUIRED:")
    print("  - Column J (Tips): Pre-filled with Gift Card tips only.")
    print("    Additional cash tips may need to be added manually.")
    print("  - Column L (Associate Pay): Set per associate per pay period")
    print("  - Column P (Color Charges): Enter manually")
    print("  - Column Q (Credit Card Amount): Auto-computed in web app (CC+GC gross_total_amount, exclusive end date)")
    print("  - Column X (Associate Fee): Auto-computed IF associate_pay is set in config")
    print("  - Footer Row (TOTAL EMP. W/DRAWL): Enter manually")


if __name__ == '__main__':
    main()
